import pandas as pd
import numpy as np
from scipy.optimize import minimize
import openpyxl
import streamlit as st
import tempfile
import os
import plotly.graph_objects as go
 
 
# =============================================================================
# FASE 2 — MARKOWITZ
# =============================================================================
 
def calcola_matrice_covarianza(rendimenti_storici_df):
    solo_rendimenti = rendimenti_storici_df.select_dtypes(include=[np.number])
    return solo_rendimenti.cov().values
 
 
def calcola_varianza_portafoglio(pesi, matrice_cov):
    return pesi @ matrice_cov @ pesi
 
 
def _pesi_iniziali_fattibili(lower_bounds, upper_bounds):
    """
    Punto di partenza che rispetta tutti i bounds e somma a 1.
    Usa water-filling: parte dai minimi e distribuisce il budget residuo
    proporzionalmente alla capacità disponibile.
    Il semplice clip+normalizzazione può violare i bounds dopo la divisione.
    """
    pesi = lower_bounds.copy().astype(float)
    budget = 1.0 - pesi.sum()
    if budget < -1e-9:
        raise ValueError(f"Somma vincoli minimi ({pesi.sum():.4f}) > 1: problema non fattibile.")
    for _ in range(len(lower_bounds) + 1):
        if budget < 1e-12:
            break
        capacita = np.maximum(upper_bounds.astype(float) - pesi, 0.0)
        tot = capacita.sum()
        if tot < 1e-12:
            break
        pesi += np.minimum(capacita, budget * capacita / tot)
        budget = 1.0 - pesi.sum()
    return np.clip(pesi, lower_bounds, upper_bounds)
 
 
def minimizza_varianza(matrice_cov, vincoli_pesi,
                       rendimento_target=None, rendimenti_attesi=None,
                       pesi_iniziali=None):
    """
    Minimizza la varianza. Accetta un punto di partenza esterno (pesi_iniziali)
    in modo che genera_frontiera_efficiente possa passare il portafoglio precedente.
    Usa ftol stretto e molte iterazioni per maggiore accuratezza.
    """
    lower_bounds = np.array([b[0] for b in vincoli_pesi])
    upper_bounds = np.array([b[1] for b in vincoli_pesi])
 
    if pesi_iniziali is None:
        pesi_iniziali = _pesi_iniziali_fattibili(lower_bounds, upper_bounds)
 
    lista_vincoli = [{"type": "eq", "fun": lambda w: np.sum(w) - 1}]
 
    if rendimento_target is not None:
        lista_vincoli.append({
            "type": "eq",
            "fun": lambda w, r=rendimento_target: w @ rendimenti_attesi - r,
        })
 
    return minimize(
        fun=lambda w: calcola_varianza_portafoglio(w, matrice_cov),
        x0=pesi_iniziali,
        method="SLSQP",
        bounds=vincoli_pesi,
        constraints=lista_vincoli,
        options={"maxiter": 2000, "ftol": 1e-12, "disp": False},
    )
 
 
def trova_rendimento_massimo(rendimenti_attesi, vincoli_pesi):
    lower_bounds = np.array([b[0] for b in vincoli_pesi])
    upper_bounds = np.array([b[1] for b in vincoli_pesi])
    pesi_iniziali = _pesi_iniziali_fattibili(lower_bounds, upper_bounds)
 
    risultato = minimize(
        fun=lambda w: -(w @ rendimenti_attesi),
        x0=pesi_iniziali,
        method="SLSQP",
        bounds=vincoli_pesi,
        constraints={"type": "eq", "fun": lambda w: np.sum(w) - 1},
        options={"maxiter": 2000, "ftol": 1e-12, "disp": False},
    )
 
    if not risultato.success:
        raise RuntimeError(f"Calcolo rendimento massimo fallito: {risultato.message}")
 
    return -risultato.fun
 
 
def _valida_portafoglio(res, rendimenti_attesi, target, lower_bounds, upper_bounds):
    """
    Restituisce True solo se success=True e tutti i vincoli sono rispettati
    entro tolleranza numerica. Impedisce di usare risultati SLSQP non validi.
    """
    TOLS = 1e-5
    if not res.success:
        return False
    w = res.x
    if abs(w.sum() - 1.0) > TOLS:
        return False
    if target is not None and abs(w @ rendimenti_attesi - target) > 1e-4:
        return False
    if np.any(w < lower_bounds - TOLS) or np.any(w > upper_bounds + TOLS):
        return False
    return True
 
 
def genera_frontiera_efficiente(rendimenti_attesi, matrice_cov, vincoli_pesi, n_portafogli=10):
    """
    Genera n_portafogli sulla frontiera efficiente di Markowitz.
 
    Logica corretta:
      P1  = GMVP calcolato senza vincolo di rendimento (usato direttamente,
             senza ri-ottimizzazione che può scivolare sul ramo inefficiente).
      P2-P10 = minimizzazione varianza con vincolo di rendimento crescente.
             Ogni ottimizzazione parte dai pesi del portafoglio precedente
             (punto di partenza migliore rispetto ai pesi iniziali fissi).
             Il risultato viene accettato solo se success=True e i vincoli
             sono soddisfatti entro tolleranza.
    """
    TOLS = 1e-5
    lower_bounds = np.array([b[0] for b in vincoli_pesi])
    upper_bounds = np.array([b[1] for b in vincoli_pesi])
    pesi_init = _pesi_iniziali_fattibili(lower_bounds, upper_bounds)
 
    # ── Step 1: Global Minimum Variance Portfolio ────────────────────────────
    res_gmvp = minimizza_varianza(matrice_cov, vincoli_pesi, pesi_iniziali=pesi_init)
    if not res_gmvp.success:
        raise RuntimeError(f"Calcolo GMVP fallito: {res_gmvp.message}")
 
    pesi_gmvp = res_gmvp.x
    r_gmvp    = float(pesi_gmvp @ rendimenti_attesi)
    vol_gmvp  = float(np.sqrt(calcola_varianza_portafoglio(pesi_gmvp, matrice_cov)))
    sh_gmvp   = r_gmvp / vol_gmvp if vol_gmvp > 0 else 0.0
 
    # ── Step 2: Rendimento massimo raggiungibile ─────────────────────────────
    r_max = trova_rendimento_massimo(rendimenti_attesi, vincoli_pesi)
    r_max = max(r_max, r_gmvp)
 
    # ── Step 3: 10 target linearmente spaziati ──────────────────────────────
    rendimenti_target = np.linspace(r_gmvp, r_max, n_portafogli)
 
    portafogli = [{
        "pesi": pesi_gmvp, "rendimento": r_gmvp,
        "volatilita": vol_gmvp, "sharpe": sh_gmvp,
    }]
 
    # ── Step 4: P2-P10 con vincolo di rendimento ─────────────────────────────
    prev_pesi = pesi_gmvp
 
    for i, target in enumerate(rendimenti_target[1:], start=2):
        trovato = False
        # Tre tentativi con punti di partenza diversi, dal migliore al più generico
        for start in [prev_pesi, pesi_init, pesi_gmvp]:
            res = minimizza_varianza(
                matrice_cov, vincoli_pesi,
                rendimento_target=target,
                rendimenti_attesi=rendimenti_attesi,
                pesi_iniziali=start.copy(),
            )
            if _valida_portafoglio(res, rendimenti_attesi, target, lower_bounds, upper_bounds):
                trovato = True
                break
 
        if not trovato:
            raise RuntimeError(
                f"Ottimizzazione fallita per P{i} (target={target*100:.4f}%). "
                "Verificare che i vincoli consentano la frontiera efficiente."
            )
 
        w   = res.x
        r   = float(w @ rendimenti_attesi)
        vol = float(np.sqrt(calcola_varianza_portafoglio(w, matrice_cov)))
        sh  = r / vol if vol > 0 else 0.0
 
        portafogli.append({"pesi": w, "rendimento": r, "volatilita": vol, "sharpe": sh})
        prev_pesi = w
 
    # ── Step 5: Verifica monotonicità (errore reale, non fix cosmetico) ───────
    vols  = np.array([p["volatilita"] for p in portafogli])
    rends = np.array([p["rendimento"] for p in portafogli])
 
    if np.any(np.diff(vols) < -TOLS):
        raise RuntimeError(
            "ERRORE FRONTIERA: volatilità non monotona. "
            "Almeno una ottimizzazione ha prodotto un punto non valido."
        )
    if np.any(np.diff(rends) < -TOLS):
        raise RuntimeError(
            "ERRORE FRONTIERA: rendimenti non monotoni. "
            "Almeno una ottimizzazione ha prodotto un punto non valido."
        )
 
    return portafogli
 
 
def salva_allocazioni_su_excel(portafogli_frontiera, nomi_asset, percorso_file):
    workbook = openpyxl.load_workbook(percorso_file, keep_vba=True)
 
    if "risultati" not in workbook.sheetnames:
        raise ValueError("Il foglio 'risultati' non esiste nel file Excel.")
 
    foglio_risultati = workbook["risultati"]
    riga_inizio = 3
    colonna_inizio = 3
    n_asset = len(nomi_asset)
    n_portafogli = len(portafogli_frontiera)
    MAX_ASSET = 18  # righe riservate nel template
 
    # Cancella le righe degli asset non più presenti (es. precedente ottimizzazione con più asset)
    for old_idx in range(n_asset, MAX_ASSET):
        for p_idx in range(n_portafogli):
            foglio_risultati.cell(
                row=riga_inizio + old_idx,
                column=colonna_inizio + p_idx,
                value=None,
            )
 
    # Scrive i pesi per gli asset effettivi
    for indice_asset in range(n_asset):
        for indice_portafoglio, portafoglio in enumerate(portafogli_frontiera):
            foglio_risultati.cell(
                row=riga_inizio + indice_asset,
                column=colonna_inizio + indice_portafoglio,
                value=round(float(portafoglio["pesi"][indice_asset]), 6),
            )
 
    # Rendimento e volatilità in posizioni fisse del template (righe 36-37)
    for i, portafoglio in enumerate(portafogli_frontiera):
        foglio_risultati.cell(row=36, column=3 + i, value=round(float(portafoglio["rendimento"]), 6))
        foglio_risultati.cell(row=37, column=3 + i, value=round(float(portafoglio["volatilita"]), 6))
 
    workbook.save(percorso_file)
    print(f"\nRisultati salvati in '{percorso_file}'")
 
 
# =============================================================================
# FASE 4 — MICHAUD-INSPIRED RESAMPLING CON BLOCK BOOTSTRAP
# =============================================================================
 
def _estrai_rendimenti_np(rendimenti_storici_df):
    solo_num = rendimenti_storici_df.select_dtypes(include=[np.number])
    # Elimina righe con NaN invece di riempirle con zero (che distorcerebbe la covarianza)
    solo_num = solo_num.dropna(how="any")
    return solo_num.values
 
 
def genera_campione_block_bootstrap(rendimenti_np, block_size=6):
    n_periodi = rendimenti_np.shape[0]
 
    if n_periodi == 0:
        raise ValueError("Non ci sono osservazioni storiche disponibili per il bootstrap.")
 
    block_size = min(block_size, n_periodi)
    blocchi = []
    n_accumulati = 0
 
    while n_accumulati < n_periodi:
        start = np.random.randint(0, n_periodi - block_size + 1)
        blocchi.append(rendimenti_np[start:start + block_size])
        n_accumulati += block_size
 
    return np.vstack(blocchi)[:n_periodi]
 
 
def _cov_da_campione_bootstrap(campione_np):
    cov = np.cov(campione_np, rowvar=False)
 
    if np.linalg.eigvalsh(cov).min() < 1e-10:
        cov += np.eye(campione_np.shape[1]) * 1e-8
 
    return cov
 
 
def _costruisci_portafoglio(pesi, rendimenti_attesi, cov_originale):
    rendimento = float(pesi @ rendimenti_attesi)
    volatilita = float(np.sqrt(calcola_varianza_portafoglio(pesi, cov_originale)))
    sharpe = rendimento / volatilita if volatilita > 0 else 0.0
    return {"pesi": pesi, "rendimento": rendimento, "volatilita": volatilita, "sharpe": sharpe}
 
 
def _ripara_frontiera_michaud(pesi_medi, rendimenti_attesi, cov_originale, n_portafogli,
                              _profondita=0):
    """
    Garantisce che la frontiera Michaud sia una vera frontiera efficiente:
    rendimento E volatilità strettamente crescenti da P1 a P10.
 
    La media dei pesi per rank può produrre punti "dominati" (stesso rischio,
    rendimento inferiore di un altro punto). Qui:
      1. Si ordinano i punti per rendimento crescente (ordine naturale dei rank).
      2. Si eliminano i punti dominati (esiste un punto successivo con
         volatilità <= alla propria: quel punto offre più rendimento a pari
         o minor rischio).
      3. Si ricostruiscono n_portafogli punti interpolando linearmente i PESI
         dei punti efficienti adiacenti su target di rendimento equispaziati.
         La combinazione convessa di due portafogli ammissibili resta
         ammissibile (vincoli box) e somma a 1, quindi i punti ricostruiti
         sono portafogli reali, non un abbellimento numerico.
    """
    TOLS = 1e-10
 
    punti = [_costruisci_portafoglio(pesi_medi[k], rendimenti_attesi, cov_originale)
             for k in range(len(pesi_medi))]
    punti.sort(key=lambda p: p["rendimento"])
 
    # ── Filtra i punti dominati: tengo j solo se nessun punto con rendimento
    #    maggiore ha volatilità minore o uguale ──────────────────────────────
    vols = np.array([p["volatilita"] for p in punti])
    efficienti = []
    min_vol_successiva = np.inf
    for j in range(len(punti) - 1, -1, -1):
        if vols[j] < min_vol_successiva - TOLS:
            efficienti.append(punti[j])
            min_vol_successiva = vols[j]
    efficienti.reverse()
 
    if len(efficienti) < 2:
        # Frontiera degenere (tutti i punti coincidono): restituisco i rank così come sono
        return punti[:n_portafogli]
 
    # ── Refill: n_portafogli target di rendimento equispaziati, pesi ottenuti
    #    per interpolazione lineare tra i punti efficienti adiacenti ──────────
    r_eff = np.array([p["rendimento"] for p in efficienti])
    target = np.linspace(r_eff[0], r_eff[-1], n_portafogli)
 
    risultato = []
    for t in target:
        j = int(np.searchsorted(r_eff, t, side="right")) - 1
        j = max(0, min(j, len(efficienti) - 2))
        r_a, r_b = r_eff[j], r_eff[j + 1]
        alpha = 0.0 if r_b - r_a < TOLS else (t - r_a) / (r_b - r_a)
        alpha = float(np.clip(alpha, 0.0, 1.0))
        pesi = (1 - alpha) * efficienti[j]["pesi"] + alpha * efficienti[j + 1]["pesi"]
        risultato.append(_costruisci_portafoglio(pesi, rendimenti_attesi, cov_originale))
 
    # ── Verifica finale: la monotonicità deve valere sui punti ricostruiti ──
    v = np.array([p["volatilita"] for p in risultato])
    r = np.array([p["rendimento"] for p in risultato])
    if (np.any(np.diff(v) < -1e-8) or np.any(np.diff(r) < -1e-8)) and _profondita < 5:
        # L'interpolazione ha creato micro-inversioni vicino al GMVP:
        # ripeto il filtro di dominanza sui punti ricostruiti (max 5 passate).
        return _ripara_frontiera_michaud(
            np.array([p["pesi"] for p in risultato]),
            rendimenti_attesi, cov_originale, n_portafogli,
            _profondita=_profondita + 1,
        )
    return risultato
 
 
def ottimizzazione_michaud_resampling(
    rendimenti_attesi,
    rendimenti_storici_df,
    vincoli_pesi,
    n_portafogli=10,
    n_simulazioni=500,
    block_size=50,
    progress_callback=None,
    fattore_annualizzazione=252,
):
    rendimenti_np = _estrai_rendimenti_np(rendimenti_storici_df)
 
    if rendimenti_np.shape[0] == 0:
        raise ValueError("Il foglio 'Elaborazione' non contiene dati numerici validi.")
 
    media_storica = rendimenti_np.mean(axis=0)
    pesi_accumulati = []
 
    for h in range(n_simulazioni):
        try:
            campione = genera_campione_block_bootstrap(rendimenti_np, block_size)
            # Covarianza bootstrap annualizzata: coerente con i rendimenti attesi annui
            cov_bootstrap = _cov_da_campione_bootstrap(campione) * fattore_annualizzazione
 
            # Rumore = errore di stima della media, derivato dallo STESSO campione
            # bootstrap e annualizzato. Prima si usava la std PERIODICA (giornaliera)
            # come rumore su rendimenti ANNUI: scala sbagliata e rumore scorrelato
            # dalla covarianza simulata. Con la scala errata il rumore domina il
            # segnale, i portafogli medi dei rank bassi collassano tutti verso la
            # stessa volatilità e alcuni risultano dominati.
            rumore = (campione.mean(axis=0) - media_storica) * fattore_annualizzazione
            rendimenti_perturbati = rendimenti_attesi + rumore
 
            portafogli_h = genera_frontiera_efficiente(
                rendimenti_perturbati, cov_bootstrap, vincoli_pesi, n_portafogli=n_portafogli,
            )
            pesi_accumulati.append(np.array([p["pesi"] for p in portafogli_h]))
 
        except Exception as e:
            print(f"Simulazione {h + 1} fallita: {e}")
 
        if progress_callback and (h + 1) % 10 == 0:
            progress_callback((h + 1) / n_simulazioni)
 
    n_riuscite = len(pesi_accumulati)
    print(f"\nSimulazioni completate: {n_riuscite}/{n_simulazioni}")
 
    if n_riuscite == 0:
        raise RuntimeError("Nessuna simulazione Michaud è riuscita. Controlla dati, vincoli e rendimenti.")
 
    pesi_medi = np.stack(pesi_accumulati, axis=0).mean(axis=0)
    pesi_medi /= pesi_medi.sum(axis=1, keepdims=True)
 
    cov_originale = calcola_matrice_covarianza(rendimenti_storici_df) * fattore_annualizzazione
 
    # NIENTE sort per volatilità: l'associazione per rank (P1 = GMVP medio,
    # P10 = max rendimento medio) va preservata. Il vecchio sort rimescolava
    # i portafogli quando le volatilità dei primi rank erano quasi identiche,
    # producendo rendimenti DECRESCENTI da P1 a P5.
    return _ripara_frontiera_michaud(pesi_medi, rendimenti_attesi, cov_originale, n_portafogli)
 
 
def salva_michaud_su_excel(portafogli_michaud, nomi_asset, percorso_file):
    workbook = openpyxl.load_workbook(percorso_file, keep_vba=True)
 
    if "Michaud" not in workbook.sheetnames:
        raise ValueError("Il foglio 'Michaud' non esiste nel file Excel.")
 
    foglio = workbook["Michaud"]
    riga_inizio = 3
    colonna_inizio = 2
    n_asset = len(nomi_asset)
    MAX_ASSET = 18  # righe riservate nel template
 
    # Cancella le righe degli asset non più presenti
    for old_idx in range(n_asset, MAX_ASSET):
        for p_idx in range(10):
            foglio.cell(
                row=riga_inizio + old_idx,
                column=colonna_inizio + p_idx,
                value=None,
            )
 
    # Scrive i pesi per gli asset effettivi
    for indice_asset in range(n_asset):
        for indice_portafoglio, portafoglio in enumerate(portafogli_michaud):
            foglio.cell(
                row=riga_inizio + indice_asset,
                column=colonna_inizio + indice_portafoglio,
                value=round(float(portafoglio["pesi"][indice_asset]), 6),
            )
 
    # Rendimento e volatilità in posizioni fisse del template (righe 27-28)
    for i, portafoglio in enumerate(portafogli_michaud):
        foglio.cell(row=27, column=2 + i, value=round(float(portafoglio["rendimento"]), 6))
        foglio.cell(row=28, column=2 + i, value=round(float(portafoglio["volatilita"]), 6))
 
    workbook.save(percorso_file)
    print(f"\nRisultati Michaud salvati in '{percorso_file}'")
 
 
# =============================================================================
# FUNZIONE PRINCIPALE
# =============================================================================
 
def _rileva_scala_rendimenti(dati_clean, status=print):
    """
    Se i valori sono in punti percentuali (es. 5.86 invece di 0.0586),
    li riporta in forma decimale. Un rendimento periodale reale non supera
    quasi mai il 100% in valore assoluto: se il 95° percentile dei valori
    assoluti è > 1, i dati sono chiaramente in punti percentuali.
    """
    q95 = float(np.nanquantile(np.abs(dati_clean.values), 0.95))
    if q95 > 1.0:
        status(
            f"ATTENZIONE: i rendimenti in 'Elaborazione' risultano in punti percentuali "
            f"(95° percentile |r| = {q95:.2f}). Divisi per 100 automaticamente."
        )
        return dati_clean / 100.0
    return dati_clean
 
 
def _estrai_colonna_date(dati_raw):
    """
    Cerca nel foglio Elaborazione una colonna interpretabile come date.
    Restituisce una Series datetime allineata all'indice del DataFrame, o None.
    """
    for col in dati_raw.columns:
        s = dati_raw[col]
        if pd.api.types.is_datetime64_any_dtype(s):
            if s.notna().mean() > 0.8:
                return s
        elif s.dtype == object:
            parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
            if parsed.notna().mean() > 0.8:
                return parsed
    return None
 
 
def _fattore_da_gap(gap_giorni):
    if gap_giorni <= 5:
        return 252, "giornaliera"
    if gap_giorni <= 10:
        return 52, "settimanale"
    if gap_giorni <= 45:
        return 12, "mensile"
    if gap_giorni <= 120:
        return 4, "trimestrale"
    return 1, "annuale"
 
 
def _regolarizza_serie(rendimenti_df, date_series, status=print):
    """
    Gestisce serie storiche con date IRREGOLARI (buchi dovuti, ad esempio,
    all'eliminazione delle righe in cui manca un asset).
 
    Se le distanze tra osservazioni sono disomogenee, ogni rendimento copre
    un orizzonte diverso (2 giorni, una settimana, un mese...) e la
    covarianza "per periodo" non è definita in modo coerente. In quel caso:
      1. ricostruisce l'indice di prezzo di ogni asset cumulando (1+r);
      2. lo ricampiona su una griglia regolare (settimanale o mensile,
         a seconda dell'ampiezza dei buchi);
      3. ricalcola i rendimenti sulla griglia regolare.
 
    Restituisce (rendimenti_df, fattore, descrizione). Se le date sono
    regolari, i dati passano invariati con il fattore dedotto dal gap mediano.
    """
    date = pd.to_datetime(date_series.values)
    gaps = pd.Series(date).diff().dt.days.dropna()
    gap_med = float(gaps.median())
    gap_max = float(gaps.max())
 
    # Irregolare solo se esistono buchi molto più ampi del passo tipico.
    # La soglia +7 assorbe weekend (gap 3 su dati giornalieri) e festività,
    # e le settimane lunghe sui dati settimanali.
    serie_irregolare = gap_max > 3 * gap_med + 7
 
    if not serie_irregolare:
        fattore, nome = _fattore_da_gap(gap_med)
        return rendimenti_df, fattore, f"{nome} (distanza mediana {gap_med:.0f} gg, regolare)"
 
    # ── Serie irregolare: ricostruzione prezzi e ricampionamento ────────────
    if gap_max <= 8:
        regola, fattore, nome = "W-FRI", 52, "settimanale"
    else:
        regola, fattore, nome = "ME", 12, "mensile"
 
    status(
        f"ATTENZIONE: date irregolari in 'Elaborazione' "
        f"(gap mediano {gap_med:.0f} gg, massimo {gap_max:.0f} gg). "
        f"Ogni rendimento copre un orizzonte diverso: i prezzi vengono "
        f"ricostruiti e ricampionati su griglia {nome}."
    )
 
    # Ricostruzione dell'indice di prezzo. Il convertitore a monte produce
    # LOG return, quindi il prezzo è exp(somma cumulata). Se i dati fossero
    # rendimenti semplici, l'approssimazione resta accurata per valori
    # periodali tipici e la serie ricampionata rimane internamente coerente.
    prezzi = np.exp(rendimenti_df.cumsum())
    prezzi.index = date
    prezzi_reg = prezzi.resample(regola).last().dropna(how="any")
    rendimenti_reg = np.log(prezzi_reg / prezzi_reg.shift(1)).dropna(how="any").reset_index(drop=True)
 
    status(
        f"Ricampionamento completato: {len(rendimenti_df)} osservazioni irregolari "
        f"→ {len(rendimenti_reg)} rendimenti {nome} regolari."
    )
 
    return rendimenti_reg, fattore, f"{nome} (ricampionata da serie irregolare)"
 
 
def _rileva_frequenza_dati(dati_clean, status=print):
    """
    Fallback quando non esiste una colonna di date: stima la frequenza
    dalla magnitudine dei rendimenti.
    """
    med_abs = float(np.nanmedian(np.abs(dati_clean.values)))
    if med_abs < 0.012:
        return 252, f"giornaliera (stimata: |r| mediano {med_abs:.2%}, nessuna colonna date trovata)"
    if med_abs < 0.025:
        return 52, f"settimanale (stimata: |r| mediano {med_abs:.2%}, nessuna colonna date trovata)"
    return 12, f"mensile (stimata: |r| mediano {med_abs:.2%}, nessuna colonna date trovata)"
 
 
def run_optimizer(excel_path, progress_callback=None, status_callback=None):
    def _s(msg):
        if status_callback:
            status_callback(msg)
        print(msg)
 
    print("\n" + "=" * 70)
    print("AVVIO OTTIMIZZATORE")
    print("=" * 70)
    print(f"File selezionato: {excel_path}")
 
    _s("Reading Excel data...")
 
    # --- Legge nomi asset e rendimenti attesi dal foglio Input (colonne B, D) ---
    datiInput_raw = pd.read_excel(
        excel_path, sheet_name="Input", usecols="B,D", header=1, nrows=20,
    ).dropna(axis=1, how="all")
    datiInput_raw = datiInput_raw.loc[:, ~datiInput_raw.columns.str.startswith("Unnamed")]
    datiInput_raw = datiInput_raw.iloc[2:]  # salta le 2 righe di sotto-intestazione
 
    if datiInput_raw.shape[1] < 2:
        raise ValueError(
            "Il foglio 'Input' non contiene abbastanza colonne in B e D. "
            "Verificare che i nomi asset siano in colonna B e i rendimenti attesi in colonna D."
        )
 
    # --- Legge i vincoli dal foglio Input (colonne N, O) ---
    vincoli_raw = pd.read_excel(
        excel_path, sheet_name="Input", usecols="N,O", header=1, nrows=20,
    )
    if vincoli_raw.shape[1] < 2:
        raise ValueError(
            "Il foglio 'Input' non contiene abbastanza colonne in N e O per i vincoli."
        )
    vincoli_raw = vincoli_raw.iloc[:, :2].copy()
    vincoli_raw.columns = ["Min", "Max"]
    vincoli_raw = vincoli_raw.iloc[2:]  # salta le 2 righe di sotto-intestazione
 
    # --- Filtra le sole righe con dati validi (scarta righe vuote) ---
    col_name   = datiInput_raw.columns[0]
    col_return = datiInput_raw.columns[1]
    datiInput = (
        datiInput_raw
        .dropna(subset=[col_name, col_return])
        .reset_index(drop=True)
    )
    vincoli = (
        vincoli_raw
        .dropna(subset=["Min", "Max"])
        .reset_index(drop=True)
    )
 
    # Allinea le due tabelle al conteggio minore in caso di leggero disallineamento
    n_asset = min(len(datiInput), len(vincoli))
    datiInput = datiInput.iloc[:n_asset].copy()
    vincoli   = vincoli.iloc[:n_asset].copy()
 
    # --- Controllo numero asset (5–18) ---
    if n_asset < 5:
        raise ValueError(
            f"Numero di asset validi insufficiente: {n_asset}. "
            f"Inserire almeno 5 asset compilati nel foglio 'Input'."
        )
    if n_asset > 18:
        raise ValueError(
            f"Numero di asset eccessivo: {n_asset}. "
            f"Il massimo consentito è 18 asset."
        )
 
    nomi_asset        = datiInput.iloc[:, 0].values
    rendimenti_attesi = datiInput.iloc[:, 1].values.astype(float)
 
    print(f"\nAsset validi trovati: {n_asset}")
    print(f"Asset: {list(nomi_asset)}")
 
    # --- Legge i rendimenti storici dal foglio Elaborazione ---
    datiElaborazione_raw = pd.read_excel(excel_path, sheet_name="Elaborazione")
    solo_num = datiElaborazione_raw.select_dtypes(include=[np.number])
 
    if solo_num.shape[1] < n_asset:
        raise ValueError(
            f"Il foglio 'Elaborazione' ha solo {solo_num.shape[1]} colonne numeriche, "
            f"ma ci sono {n_asset} asset. Verificare la struttura del foglio."
        )
 
    # Prende solo le prime n_asset colonne numeriche per allinearsi agli asset letti
    datiElaborazione = solo_num.iloc[:, :n_asset].copy()
    datiElaborazione.columns = list(nomi_asset)
 
    # Segnala valori mancanti e li gestisce eliminando le righe incomplete
    n_nan = datiElaborazione.isna().sum().sum()
    if n_nan > 0:
        print(
            f"\nATTENZIONE: trovati {n_nan} valori mancanti in 'Elaborazione'. "
            f"Le righe incomplete verranno escluse dal calcolo."
        )
 
    # Allinea le date (se presenti) alle righe dei rendimenti PRIMA del dropna,
    # così restano sincronizzate con le osservazioni superstiti.
    _date_col = _estrai_colonna_date(datiElaborazione_raw)
    _mask_valide = datiElaborazione.notna().all(axis=1)
    if _date_col is not None:
        _date_col = _date_col.reindex(datiElaborazione.index)
        _mask_valide &= _date_col.notna()
 
    datiElaborazione_clean = datiElaborazione[_mask_valide].reset_index(drop=True)
 
    # --- Rilevamento automatico di scala e frequenza dei dati ---
    datiElaborazione_clean = _rileva_scala_rendimenti(datiElaborazione_clean, status=print)
 
    if _date_col is not None:
        _date_clean = _date_col[_mask_valide].reset_index(drop=True)
        datiElaborazione_clean, fattore_annualizzazione, freq_desc = _regolarizza_serie(
            datiElaborazione_clean, _date_clean, status=print,
        )
    else:
        fattore_annualizzazione, freq_desc = _rileva_frequenza_dati(
            datiElaborazione_clean, status=print,
        )
 
    _s(f"Data frequency: {freq_desc} — annualization ×{fattore_annualizzazione}")
    print(f"\nFrequenza dati rilevata: {freq_desc} → fattore di annualizzazione {fattore_annualizzazione}")
 
    min_obs_needed = n_asset + 1
    if len(datiElaborazione_clean) < min_obs_needed:
        raise ValueError(
            f"Dati storici insufficienti: solo {len(datiElaborazione_clean)} righe complete "
            f"su {len(datiElaborazione)} totali. Servono almeno {min_obs_needed} osservazioni."
        )
 
    # --- Validazione vincoli ---
    lower_bounds = vincoli["Min"].values.astype(float)
    upper_bounds = vincoli["Max"].values.astype(float)
 
    for i in range(n_asset):
        if lower_bounds[i] > upper_bounds[i] + 1e-9:
            raise ValueError(
                f"Vincolo errato per l'asset '{nomi_asset[i]}' (posizione {i + 1}): "
                f"min ({lower_bounds[i]:.4f}) > max ({upper_bounds[i]:.4f})."
            )
 
    if lower_bounds.sum() > 1.0 + 1e-6:
        raise ValueError(
            f"La somma dei vincoli minimi ({lower_bounds.sum():.4f}) supera 1. "
            f"Nessun portafoglio è fattibile con questi vincoli."
        )
    if upper_bounds.sum() < 1.0 - 1e-6:
        raise ValueError(
            f"La somma dei vincoli massimi ({upper_bounds.sum():.4f}) è inferiore a 1. "
            f"Nessun portafoglio è fattibile con questi vincoli."
        )
 
    print(f"\nSomma minimi: {lower_bounds.sum():.4f} — Somma massimi: {upper_bounds.sum():.4f}")
 
    vincoli_pesi = list(zip(lower_bounds, upper_bounds))
 
    # --- Matrice di covarianza annualizzata con il fattore rilevato ---
    _s("Calculating covariance matrix...")
    matrice_covarianza = datiElaborazione_clean.cov().values * fattore_annualizzazione
 
    if matrice_covarianza.shape != (n_asset, n_asset):
        raise RuntimeError(
            f"Errore interno: matrice di covarianza {matrice_covarianza.shape} "
            f"non corrisponde a {n_asset} asset."
        )
 
    # --- Markowitz ---
    _s("Generating Markowitz portfolios...")
    portafogli_frontiera = genera_frontiera_efficiente(
        rendimenti_attesi, matrice_covarianza, vincoli_pesi, n_portafogli=10,
    )
    salva_allocazioni_su_excel(portafogli_frontiera, nomi_asset, excel_path)
 
    # --- Michaud ---
    # Block size proporzionato alla frequenza: ~un trimestre di osservazioni
    block_size = {252: 50, 52: 13, 12: 6, 4: 2}.get(fattore_annualizzazione, 6)
    _s("Running 500 Michaud simulations...")
    portafogli_michaud = ottimizzazione_michaud_resampling(
        rendimenti_attesi=rendimenti_attesi,
        rendimenti_storici_df=datiElaborazione_clean,
        vincoli_pesi=vincoli_pesi,
        n_portafogli=10,
        n_simulazioni=500,
        block_size=block_size,
        progress_callback=progress_callback,
        fattore_annualizzazione=fattore_annualizzazione,
    )
 
    _s("Preparing results...")
    salva_michaud_su_excel(portafogli_michaud, nomi_asset, excel_path)
 
    print("\n" + "=" * 70)
    print("OTTIMIZZAZIONE COMPLETATA")
    print("=" * 70)
 
    return portafogli_frontiera, portafogli_michaud, nomi_asset
 
 
# =============================================================================
# STREAMLIT — CSS
# =============================================================================
 
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Space+Grotesk:wght@500;600;700&display=swap');
 
#MainMenu, footer, header,
[data-testid="stDecoration"],
[data-testid="stStatusWidget"] { display: none !important; }
 
:root {
    --bg:        #050505;
    --panel:     #0D0D0F;
    --panel-2:   #141417;
    --border:    #24242A;
    --text:      #F7F7F8;
    --text-2:    #9A9AA3;
    --accent:    #4C7DFF;
    --accent-h:  #6B93FF;
    --success:   #22C55E;
    --danger:    #F87171;
}
 
html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}
 
.stApp, [data-testid="stAppViewContainer"] > .main { background: var(--bg); }
 
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--panel); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
 
h1, h2, h3, h4 { color: var(--text) !important; }
 
*:focus-visible { outline: 2px solid var(--accent-h); outline-offset: 2px; }
 
@media (prefers-reduced-motion: reduce) {
    *, *::before, *::after { animation: none !important; transition: none !important; }
}
 
/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: var(--panel);
    border-right: 1px solid var(--border);
}
[data-testid="stSidebar"] > div { padding-top: 1.2rem; }
 
.brand {
    display: flex; align-items: center; gap: 12px;
    padding: 4px 4px 18px;
    border-bottom: 1px solid var(--border);
    margin-bottom: 14px;
}
.brand-mark {
    width: 34px; height: 34px; border-radius: 9px;
    background: var(--accent);
    display: flex; align-items: center; justify-content: center;
    font-size: 1rem; color: #ffffff; flex-shrink: 0;
    font-family: 'Space Grotesk', sans-serif; font-weight: 700;
}
.brand-name {
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600; font-size: 1rem; color: var(--text);
    line-height: 1.2; letter-spacing: -0.01em;
}
.brand-tag { font-size: 0.68rem; color: var(--text-2); margin-top: 1px; }
 
/* radio della sidebar reso come nav */
[data-testid="stSidebar"] [role="radiogroup"] { gap: 4px; }
[data-testid="stSidebar"] [role="radiogroup"] > label {
    background: transparent;
    border-radius: 10px;
    padding: 10px 12px;
    margin: 0;
    width: 100%;
    color: var(--text-2);
    font-size: 0.9rem; font-weight: 500;
    transition: background 0.15s ease, color 0.15s ease;
    cursor: pointer;
}
[data-testid="stSidebar"] [role="radiogroup"] > label:hover {
    background: var(--panel-2);
    color: var(--text);
}
[data-testid="stSidebar"] [role="radiogroup"] > label > div:first-child { display: none; }
[data-testid="stSidebar"] [role="radiogroup"] > label:has(input:checked) {
    background: rgba(76, 125, 255, 0.12);
    color: var(--accent);
    font-weight: 600;
}
[data-testid="stSidebar"] [role="radiogroup"] > label:has(input:checked) p { color: var(--accent) !important; }
[data-testid="stSidebar"] [role="radiogroup"] p { font-size: 0.9rem; color: inherit !important; }
 
.side-status {
    margin-top: 18px; padding: 12px;
    background: var(--panel-2);
    border: 1px solid var(--border);
    border-radius: 12px;
    font-size: 0.78rem; color: var(--text-2);
    display: flex; align-items: center; gap: 8px;
}
.side-foot {
    margin-top: 22px; padding-top: 14px;
    border-top: 1px solid var(--border);
    font-size: 0.72rem; color: var(--text-2); line-height: 1.7;
}
.side-foot a { color: var(--accent); text-decoration: none; }
.side-foot a:hover { color: var(--accent-h); }
 
/* ── Header pagina ── */
.hdr-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2.1rem; font-weight: 700; color: var(--text);
    line-height: 1.15; margin: 0 0 6px; letter-spacing: -0.02em;
}
.hdr-sub { font-size: 0.95rem; color: var(--text-2); margin-bottom: 4px; }
 
/* ── Section label ── */
.sec-label {
    font-size: 0.68rem; font-weight: 600;
    letter-spacing: 0.1em; text-transform: uppercase;
    color: var(--text-2);
    margin-bottom: 14px; padding-bottom: 10px;
    border-bottom: 1px solid var(--border);
}
 
/* ── Metric card ── */
.mcard {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 18px 14px;
    text-align: center;
    transition: border-color 0.2s ease;
}
.mcard:hover { border-color: var(--accent); }
.mcard-label {
    font-size: 0.66rem; color: var(--text-2);
    text-transform: uppercase; letter-spacing: 0.08em;
    margin-bottom: 8px; font-weight: 500;
}
.mcard-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.5rem; font-weight: 600; color: var(--text);
    line-height: 1; font-variant-numeric: tabular-nums;
}
 
/* ── File card ── */
.file-card {
    display: flex; align-items: center; gap: 14px;
    background: var(--panel);
    border: 1px solid var(--border);
    border-left: 3px solid var(--success);
    border-radius: 12px;
    padding: 14px 18px;
}
.file-name { font-weight: 600; color: var(--text); font-size: 0.92rem; }
.file-meta { color: var(--text-2); font-size: 0.78rem; margin-top: 2px; }
 
/* ── Status pill ── */
.spill {
    display: inline-flex; align-items: center; gap: 8px;
    background: var(--panel-2);
    border: 1px solid var(--border);
    border-radius: 999px;
    padding: 7px 16px;
    font-size: 0.82rem; color: var(--text-2); font-weight: 500;
}
.spill--active  { border-color: rgba(76,125,255,0.4);  color: var(--accent); }
.spill--ok      { border-color: rgba(34,197,94,0.35);  color: var(--success); }
.spill--err     { border-color: rgba(248,113,113,0.35); color: var(--danger); }
.spill-dot { width: 7px; height: 7px; border-radius: 50%; background: var(--text-2); flex-shrink: 0; }
.spill--active .spill-dot { background: var(--accent); animation: pulse 1.5s infinite; }
.spill--ok  .spill-dot { background: var(--success); }
.spill--err .spill-dot { background: var(--danger); }
 
/* ── Compare card ── */
.ccard {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 22px;
}
.ccard-title {
    font-size: 0.67rem; font-weight: 700;
    letter-spacing: 0.1em; text-transform: uppercase;
    color: var(--text-2);
    margin-bottom: 14px; padding-bottom: 10px;
    border-bottom: 1px solid var(--border);
    display: grid; grid-template-columns: 1fr 1fr;
}
.crow {
    display: flex; justify-content: space-between; align-items: center;
    padding: 7px 0;
    border-bottom: 1px solid var(--border);
    font-size: 0.86rem;
}
.crow:last-child { border-bottom: none; }
.crow-key { color: var(--text-2); }
.crow-val { font-weight: 600; color: var(--text); font-variant-numeric: tabular-nums; }
.crow-better { color: var(--success) !important; }
.crow-worse  { color: var(--danger)  !important; }
 
/* ── Buttons ── */
div.stButton > button {
    background: var(--accent);
    color: #ffffff !important;
    border: none !important;
    border-radius: 12px;
    padding: 0.65rem 2rem;
    font-size: 0.93rem; font-weight: 600;
    transition: background 0.15s ease;
    width: 100%;
}
div.stButton > button:hover:not(:disabled) { background: var(--accent-h); }
div.stButton > button:disabled {
    background: var(--panel-2) !important;
    color: #4B4B55 !important;
}
 
[data-testid="stDownloadButton"] > button {
    background: var(--accent) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 12px !important;
    font-size: 1rem !important; font-weight: 600 !important;
    padding: 0.75rem 2rem !important;
    transition: background 0.15s ease !important;
    width: 100% !important;
}
[data-testid="stDownloadButton"] > button:hover { background: var(--accent-h) !important; }
 
/* ── File uploader: stili nel blocco CSS_PORTOPT ── */
 
/* ── Progress bar ── */
[data-testid="stProgressBar"] > div { background: var(--panel-2); border-radius: 999px; height: 6px; }
[data-testid="stProgressBar"] > div > div { background: var(--accent); border-radius: 999px; }
 
/* ── Tabs ── */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 4px; gap: 4px;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    background: transparent !important;
    color: var(--text-2) !important;
    border-radius: 8px;
    padding: 8px 20px;
    font-weight: 500; border: none !important; font-size: 0.87rem;
}
[data-testid="stTabs"] [aria-selected="true"] {
    background: rgba(76,125,255,0.12) !important;
    color: var(--accent) !important;
}
 
/* ── Expander ── */
[data-testid="stExpander"] {
    background: var(--panel);
    border: 1px solid var(--border) !important;
    border-radius: 12px;
}
[data-testid="stExpander"] > details > summary {
    color: var(--text) !important;
    font-weight: 500; font-size: 0.9rem;
}
 
/* ── Selectbox ── */
[data-testid="stSelectbox"] > div > div {
    background: var(--panel-2) !important;
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
    color: var(--text) !important;
}
 
[data-testid="stAlert"] { border-radius: 10px; }
 
hr { border-color: var(--border) !important; margin: 2rem 0 !important; }
 
/* ── Empty state ── */
.empty-card {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 48px 24px;
    text-align: center;
    color: var(--text-2);
}
.empty-card .big { font-size: 2rem; margin-bottom: 10px; }
.empty-card .title { color: var(--text); font-weight: 600; font-size: 1rem; margin-bottom: 4px; }
 
@keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.3; } }
</style>
"""
 
# =============================================================================
# STREAMLIT — HELPERS
# =============================================================================
 
C_MK = "#4C7DFF"   # accent — Markowitz
C_MI = "#9A9AA3"   # gray  — Michaud (duotone blu/grigio)
C_BG = "#050505"
C_GRID = "rgba(255,255,255,0.05)"
C_TXT2 = "#9A9AA3"
 
 
def _effective_n(pesi):
    hhi = np.sum(np.array(pesi) ** 2)
    return 1.0 / hhi if hhi > 0 else 1.0
 
 
def _best_sharpe_p(portafogli):
    return max(portafogli, key=lambda p: p["sharpe"])
 
 
def _min_vol_p(portafogli):
    return min(portafogli, key=lambda p: p["volatilita"])
 
 
def _mcard(label, value, accent=False):
    color = C_MK if accent else "#F7F7F8"
    return (
        f'<div class="mcard">'
        f'<div class="mcard-label">{label}</div>'
        f'<div class="mcard-value" style="color:{color};">{value}</div>'
        f'</div>'
    )
 
 
def _spill(text, kind=""):
    cls = f"spill spill--{kind}" if kind else "spill"
    return f'<div class="{cls}"><div class="spill-dot"></div>{text}</div>'
 
 
def _layout_grafico(fig, height, xtitle=None, ytitle=None, legenda=True):
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color=C_TXT2),
        xaxis=dict(
            title=xtitle, gridcolor=C_GRID, zeroline=False, showline=False,
            tickfont=dict(color=C_TXT2), title_font=dict(color=C_TXT2, size=11),
        ),
        yaxis=dict(
            title=ytitle, gridcolor=C_GRID, zeroline=False, showline=False,
            tickfont=dict(color=C_TXT2), title_font=dict(color=C_TXT2, size=11),
        ),
        legend=dict(
            bgcolor="#0D0D0F", bordercolor="#24242A", borderwidth=1,
            font=dict(color="#F7F7F8", size=12),
            orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
        ) if legenda else None,
        showlegend=legenda,
        hovermode="closest", margin=dict(l=20, r=20, t=55 if legenda else 20, b=20),
        height=height,
    )
    return fig
 
 
def _build_frontier_chart(pf, pm):
    fig = go.Figure()
 
    fig.add_trace(go.Scatter(
        x=[p["volatilita"] * 100 for p in pf],
        y=[p["rendimento"]  * 100 for p in pf],
        mode="lines+markers", name="Markowitz",
        line=dict(color=C_MK, width=2.5, shape="linear"),
        marker=dict(size=8, color=C_MK, line=dict(color=C_BG, width=1.5)),
        customdata=[[f"P{i+1}", p["sharpe"]] for i, p in enumerate(pf)],
        hovertemplate=(
            "<b>%{customdata[0]} — Markowitz</b><br>"
            "Volatility: %{x:.2f}%<br>Return: %{y:.2f}%<br>"
            "Sharpe: %{customdata[1]:.4f}<extra></extra>"
        ),
    ))
 
    fig.add_trace(go.Scatter(
        x=[p["volatilita"] * 100 for p in pm],
        y=[p["rendimento"]  * 100 for p in pm],
        mode="lines+markers", name="Michaud",
        line=dict(color=C_MI, width=2.5, shape="spline", smoothing=0.7),
        marker=dict(size=8, color=C_MI, line=dict(color=C_BG, width=1.5)),
        customdata=[[f"P{i+1}", p["sharpe"]] for i, p in enumerate(pm)],
        hovertemplate=(
            "<b>%{customdata[0]} — Michaud</b><br>"
            "Volatility: %{x:.2f}%<br>Return: %{y:.2f}%<br>"
            "Sharpe: %{customdata[1]:.4f}<extra></extra>"
        ),
    ))
 
    mv_mk = _min_vol_p(pf);  mv_mi = _min_vol_p(pm)
    ms_mk = _best_sharpe_p(pf); ms_mi = _best_sharpe_p(pm)
 
    fig.add_trace(go.Scatter(
        x=[mv_mk["volatilita"] * 100, mv_mi["volatilita"] * 100],
        y=[mv_mk["rendimento"]  * 100, mv_mi["rendimento"]  * 100],
        mode="markers", name="Min Volatility",
        marker=dict(size=13, color="#F7F7F8", symbol="diamond", line=dict(color=C_MK, width=2)),
        hovertemplate="<b>Min Volatility</b><br>Vol: %{x:.2f}%<br>Return: %{y:.2f}%<extra></extra>",
    ))
 
    fig.add_trace(go.Scatter(
        x=[ms_mk["volatilita"] * 100, ms_mi["volatilita"] * 100],
        y=[ms_mk["rendimento"]  * 100, ms_mi["rendimento"]  * 100],
        mode="markers", name="Max Sharpe",
        marker=dict(size=13, color="#22C55E", symbol="star", line=dict(color=C_BG, width=1)),
        hovertemplate="<b>Max Sharpe</b><br>Vol: %{x:.2f}%<br>Return: %{y:.2f}%<extra></extra>",
    ))
 
    return _layout_grafico(fig, 490, "Annualized Volatility (%)", "Expected Annual Return (%)")
 
 
def _build_single_chart(portafogli, color, name, line_shape="linear"):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=[p["volatilita"] * 100 for p in portafogli],
        y=[p["rendimento"]  * 100 for p in portafogli],
        mode="lines+markers", name=name,
        line=dict(color=color, width=2.5, shape=line_shape),
        marker=dict(size=9, color=color, line=dict(color=C_BG, width=1.5)),
        customdata=[[f"P{i+1}", p["sharpe"]] for i, p in enumerate(portafogli)],
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Vol: %{x:.2f}%<br>Return: %{y:.2f}%<br>"
            "Sharpe: %{customdata[1]:.4f}<extra></extra>"
        ),
    ))
    return _layout_grafico(fig, 320, "Volatility (%)", "Return (%)", legenda=False)
 
 
def _build_alloc_chart(pf, pm, nomi_asset, idx):
    pesi_mk = pf[idx]["pesi"] * 100
    pesi_mi = pm[idx]["pesi"] * 100
    labels  = [str(n) for n in nomi_asset]
 
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Markowitz", x=labels, y=list(pesi_mk), marker_color=C_MK,
        text=[f"{v:.1f}%" for v in pesi_mk], textposition="outside",
        textfont=dict(color=C_TXT2, size=9),
        hovertemplate="<b>Markowitz</b><br>%{x}: %{y:.2f}%<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Michaud", x=labels, y=list(pesi_mi), marker_color=C_MI,
        text=[f"{v:.1f}%" for v in pesi_mi], textposition="outside",
        textfont=dict(color=C_TXT2, size=9),
        hovertemplate="<b>Michaud</b><br>%{x}: %{y:.2f}%<extra></extra>",
    ))
    fig.update_layout(barmode="group", bargap=0.18, bargroupgap=0.06)
    fig = _layout_grafico(fig, 400, None, "Weight (%)")
    fig.update_layout(
        xaxis=dict(gridcolor=C_GRID, zeroline=False,
                   tickfont=dict(color=C_TXT2, size=9), tickangle=-30),
        legend=dict(bgcolor="#0D0D0F", bordercolor="#24242A", borderwidth=1,
                    font=dict(color="#F7F7F8")),
        hovermode="x unified", margin=dict(l=20, r=20, t=20, b=70),
    )
    return fig
 
 
def _portfolio_df(portafogli, nomi_asset):
    rows = []
    for i, p in enumerate(portafogli):
        row = {
            "#": f"P{i+1}",
            "Return":     f"{p['rendimento']:.2%}",
            "Volatility": f"{p['volatilita']:.2%}",
            "Sharpe":     f"{p['sharpe']:.4f}",
        }
        for j, n in enumerate(nomi_asset):
            row[str(n)] = f"{p['pesi'][j]:.1%}"
        rows.append(row)
    return pd.DataFrame(rows)
 
 
def _styled_table(portafogli, nomi_asset, idx_sharpe, idx_vol):
    df = _portfolio_df(portafogli, nomi_asset)
 
    def hl(row):
        i = int(row["#"][1:]) - 1
        if i == idx_sharpe:
            return ["background-color: rgba(76,125,255,0.16); color: #F7F7F8"] * len(row)
        if i == idx_vol:
            return ["background-color: rgba(255,255,255,0.06); color: #F7F7F8"] * len(row)
        return [""] * len(row)
 
    return df.style.apply(hl, axis=1)
 
 
def _compare_cards(pf, pm):
    bs_mk = _best_sharpe_p(pf)
    bs_mi = _best_sharpe_p(pm)
    eff_mk = _effective_n(bs_mk["pesi"])
    eff_mi = _effective_n(bs_mi["pesi"])
    conc_mk = max(bs_mk["pesi"]) * 100
    conc_mi = max(bs_mi["pesi"]) * 100
 
    def _cls(val_mi, val_mk, higher_better):
        better = val_mi > val_mk if higher_better else val_mi < val_mk
        return "crow-better" if better else "crow-worse"
 
    def _row(key, mk_str, mi_str, mi_cls):
        return (
            f'<div class="crow">'
            f'<span class="crow-key">{key}</span>'
            f'<span class="crow-val">{mk_str}</span>'
            f'<span class="crow-val {mi_cls}">{mi_str}</span>'
            f'</div>'
        )
 
    ret_cls  = _cls(bs_mi["rendimento"], bs_mk["rendimento"], True)
    vol_cls  = _cls(bs_mi["volatilita"], bs_mk["volatilita"], False)
    shr_cls  = _cls(bs_mi["sharpe"],     bs_mk["sharpe"],     True)
    conc_cls = _cls(conc_mi, conc_mk, False)
    eff_cls  = _cls(eff_mi,  eff_mk,  True)
 
    return (
        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:4px;">'
 
        f'<div class="ccard" style="border-top:2px solid {C_MK};">'
        f'<div class="ccard-title"><span>Metric</span><span style="color:{C_MK};">Markowitz vs Michaud</span></div>'
        + _row("Expected Return",  f"{bs_mk['rendimento']:.2%}", f"{bs_mi['rendimento']:.2%}", ret_cls)
        + _row("Volatility",       f"{bs_mk['volatilita']:.2%}", f"{bs_mi['volatilita']:.2%}", vol_cls)
        + _row("Sharpe Ratio",     f"{bs_mk['sharpe']:.4f}",     f"{bs_mi['sharpe']:.4f}",     shr_cls)
        + _row("Largest Position", f"{conc_mk:.1f}%",            f"{conc_mi:.1f}%",            conc_cls)
        + _row("Effective Assets", f"{eff_mk:.1f}",              f"{eff_mi:.1f}",              eff_cls)
        + '</div>'
 
        f'<div class="ccard" style="border-top:2px solid {C_MI};">'
        f'<div class="ccard-title"><span>Metric</span><span style="color:{C_MI};">Max Sharpe Portfolio</span></div>'
        + _row("Expected Return",  f"{bs_mk['rendimento']:.2%}", f"{bs_mi['rendimento']:.2%}", ret_cls)
        + _row("Volatility",       f"{bs_mk['volatilita']:.2%}", f"{bs_mi['volatilita']:.2%}", vol_cls)
        + _row("Sharpe Ratio",     f"{bs_mk['sharpe']:.4f}",     f"{bs_mi['sharpe']:.4f}",     shr_cls)
        + _row("Largest Position", f"{conc_mk:.1f}%",            f"{conc_mi:.1f}%",            conc_cls)
        + _row("Effective Assets", f"{eff_mk:.1f}",              f"{eff_mi:.1f}",              eff_cls)
        + '</div>'
 
        '</div>'
        '<p style="font-size:0.72rem;color:#9A9AA3;margin-top:8px;text-align:right;">'
        'Green = Michaud improved vs Markowitz &nbsp;|&nbsp; Red = Michaud underperformed on this metric'
        '</p>'
    )
 
 
# =============================================================================
# STREAMLIT — PAGINE
# =============================================================================
 
# Icone SVG (line style) usate come mask così ereditano il colore del testo.
_ICO = {
    "upload":  "data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='black' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M4 14.9A7 7 0 1 1 15.7 8h1.8a4.5 4.5 0 0 1 2.5 8.2'/><path d='M12 12v9'/><path d='m8 17 4-4 4 4'/></svg>",
    "results": "data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='black' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M3 3v16a2 2 0 0 0 2 2h16'/><path d='m7 13 3-3 4 4 5-6'/></svg>",
    "about":   "data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='black' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='12' cy='12' r='10'/><path d='M12 16v-4'/><path d='M12 8h.01'/></svg>",
}
 
CSS_PORTOPT = f"""
<style>
/* ── Icone nella nav della sidebar ── */
[data-testid="stSidebar"] [role="radiogroup"] > label p::before {{
    content: "";
    display: inline-block;
    width: 17px; height: 17px;
    margin-right: 11px;
    vertical-align: -3px;
    background-color: currentColor;
    -webkit-mask: var(--nav-ico) center / contain no-repeat;
    mask: var(--nav-ico) center / contain no-repeat;
}}
[data-testid="stSidebar"] [role="radiogroup"] > label:nth-of-type(1) p::before {{ --nav-ico: url("{_ICO['upload']}"); }}
[data-testid="stSidebar"] [role="radiogroup"] > label:nth-of-type(2) p::before {{ --nav-ico: url("{_ICO['results']}"); }}
[data-testid="stSidebar"] [role="radiogroup"] > label:nth-of-type(3) p::before {{ --nav-ico: url("{_ICO['about']}"); }}
 
/* ── Brand mark a barre ── */
.brand-mark {{
    background: transparent !important;
    border: none;
    width: 30px; height: 30px;
    display: flex; align-items: flex-end; gap: 3px;
}}
.brand-mark span {{
    width: 5px; border-radius: 2px; background: var(--accent);
    display: inline-block;
}}
.brand-name {{ font-size: 1.12rem; }}
 
/* ── Card informative della sidebar ── */
.side-card {{
    background: var(--panel-2);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 14px;
    margin-top: 16px;
    font-size: 0.76rem; color: var(--text-2); line-height: 1.55;
}}
.side-card .t {{ color: var(--text); font-weight: 600; font-size: 0.82rem; margin-bottom: 6px; }}
.vpill {{
    display: inline-block; margin-top: 10px;
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 999px; padding: 3px 12px;
    font-size: 0.7rem; color: var(--text-2);
}}
.side-row {{
    display: flex; align-items: center; justify-content: space-between;
    background: var(--panel-2); border: 1px solid var(--border);
    border-radius: 12px; padding: 11px 14px; margin-top: 10px;
    font-size: 0.8rem; color: var(--text);
}}
.tgl {{
    width: 36px; height: 20px; border-radius: 999px;
    background: var(--accent); position: relative; flex-shrink: 0;
}}
.tgl::after {{
    content: ""; position: absolute; right: 3px; top: 3px;
    width: 14px; height: 14px; border-radius: 50%; background: #fff;
}}
.ok-txt {{ color: var(--success); font-weight: 600; }}
 
/* ── Header pagina con tile e pill di stato ── */
.pg-head {{
    display: flex; align-items: flex-start; justify-content: space-between;
    gap: 16px; padding: 1.1rem 0 1.5rem;
}}
.pg-head-left {{ display: flex; align-items: center; gap: 20px; }}
.app-tile {{
    width: 76px; height: 76px; border-radius: 20px;
    background: var(--panel-2);
    border: 1px solid var(--border);
    box-shadow: 0 0 22px rgba(76,125,255,0.10);
    display: flex; align-items: flex-end; justify-content: center;
    gap: 5px; padding-bottom: 20px; flex-shrink: 0;
}}
.app-tile span {{ width: 7px; border-radius: 3px; background: var(--accent); display: inline-block; }}
.hdr-title {{ font-size: 2.3rem; }}
 
/* ── Drop zone: restyling ADDITIVO dell'uploader nativo ──
   Nessun display:none / visibility:hidden e nessuna sostituzione di testi:
   regole solo cosmetiche, così click e drag-and-drop non possono rompersi
   al variare della struttura interna di Streamlit. */
[data-testid="stFileUploader"] {{
    background: transparent;
    border: none;
    padding: 0;
}}
[data-testid="stFileUploaderDropzone"],
[data-testid="stFileUploader"] section {{
    flex-direction: column;
    justify-content: center;
    gap: 4px;
    min-height: 300px;
    background: rgba(13,13,15,0.65) !important;
    border: 1.5px dashed rgba(76,125,255,0.55) !important;
    border-radius: 14px !important;
    padding: 40px 24px !important;
    transition: border-color .2s ease, background .2s ease;
    cursor: pointer;
}}
[data-testid="stFileUploaderDropzone"]:hover,
[data-testid="stFileUploader"] section:hover {{
    border-color: var(--accent-h) !important;
    background: var(--panel-2) !important;
}}
[data-testid="stFileUploaderDropzoneInstructions"] {{
    display: flex; flex-direction: column; align-items: center;
    text-align: center; margin: 0;
}}
/* icona nativa: ingrandita, blu, dentro un cerchio scuro */
[data-testid="stFileUploaderDropzone"] svg {{
    width: 88px !important; height: 88px !important;
    padding: 26px; box-sizing: border-box;
    background: var(--panel-2);
    border-radius: 50%;
    color: var(--accent) !important;
    fill: var(--accent) !important;
    margin-bottom: 14px;
}}
/* testi nativi: ingranditi e ricolorati (il wording resta quello di Streamlit) */
[data-testid="stFileUploaderDropzoneInstructions"] span {{
    font-size: 1.18rem !important; font-weight: 700 !important;
    color: var(--text) !important;
}}
[data-testid="stFileUploaderDropzoneInstructions"] small {{
    font-size: 0.85rem !important;
    color: var(--text-2) !important;
}}
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] section button {{
    background: var(--accent) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    padding: 0.6rem 1.6rem !important;
    margin-top: 12px;
    transition: background .15s ease;
}}
[data-testid="stFileUploaderDropzone"] button:hover,
[data-testid="stFileUploader"] section button:hover {{ background: var(--accent-h) !important; }}
/* riga del file caricato: icone e ✕ restano native, solo colori coerenti */
[data-testid="stFileUploaderFile"] {{ color: var(--text) !important; }}
[data-testid="stFileUploaderFile"] small {{ color: var(--text-2) !important; }}
.dz-hint {{
    text-align: center; color: var(--text-2); font-size: 0.8rem;
    margin-top: 10px;
    pointer-events: none;
}}
 
/* ── Banner informativo ── */
.info-banner {{
    display: flex; align-items: center; gap: 16px;
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 16px 20px;
    margin-top: 18px;
}}
.info-banner .l1 {{ color: var(--text); font-size: 0.9rem; font-weight: 500; }}
.info-banner .l2 {{ color: var(--text-2); font-size: 0.82rem; margin-top: 2px; }}
.ib-ico {{
    width: 34px; height: 34px; border-radius: 8px; flex-shrink: 0;
    background-color: rgba(76,125,255,0.12);
    background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234C7DFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><path d='M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z'/><path d='M14 2v6h6'/><path d='M16 13H8'/><path d='M16 17H8'/></svg>");
    background-repeat: no-repeat; background-position: center; background-size: 18px;
}}
 
/* ── Riga delle feature ── */
.feat-row {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(210px, 1fr));
    gap: 10px;
    border-top: 1px solid var(--border);
    margin-top: 26px; padding-top: 22px;
}}
.feat {{ display: flex; align-items: center; gap: 12px; }}
.feat-ico {{
    width: 44px; height: 44px; border-radius: 50%; flex-shrink: 0;
    border: 1px solid rgba(76,125,255,0.45);
    background-color: rgba(76,125,255,0.08);
    background-repeat: no-repeat; background-position: center; background-size: 19px;
}}
.feat .t {{ color: var(--text); font-size: 0.86rem; font-weight: 600; }}
.feat .s {{ color: var(--text-2); font-size: 0.74rem; margin-top: 1px; }}
 
/* ── Pill "Need help?" ── */
.help-pill {{
    position: fixed; right: 26px; bottom: 22px; z-index: 999;
    background: var(--panel-2); border: 1px solid var(--border);
    border-radius: 999px; padding: 10px 18px;
    color: var(--text) !important; font-size: 0.85rem; font-weight: 500;
    text-decoration: none !important;
    transition: border-color .15s ease;
}}
.help-pill:hover {{ border-color: var(--accent); color: var(--accent) !important; }}
</style>
"""
 
 
def _pill_stato():
    if st.session_state.get("results"):
        return '<div class="spill spill--ok"><div class="spill-dot"></div>Results ready</div>'
    return '<div class="spill spill--ok"><div class="spill-dot"></div>Ready</div>'
 
 
def _header_pagina(titolo, sottotitolo):
    st.markdown(
        f'<div class="pg-head">'
        f'<div class="pg-head-left">'
        f'<div class="app-tile"><span style="height:14px;"></span><span style="height:24px;"></span>'
        f'<span style="height:19px;"></span><span style="height:30px;"></span></div>'
        f'<div><div class="hdr-title">{titolo}</div>'
        f'<div class="hdr-sub">{sottotitolo}</div></div>'
        f'</div>'
        f'{_pill_stato()}'
        f'</div>',
        unsafe_allow_html=True,
    )
 
def _sidebar():
    with st.sidebar:
        st.markdown(
            '<div class="brand">'
            '<div class="brand-mark">'
            '<span style="height:11px;"></span><span style="height:19px;"></span>'
            '<span style="height:15px;"></span><span style="height:24px;"></span>'
            '</div>'
            '<div><div class="brand-name">PortOpt</div></div>'
            '</div>',
            unsafe_allow_html=True,
        )
 
        # navigazione differita: impostata prima di creare il widget radio
        goto = st.session_state.pop("_nav_goto", None)
        if goto:
            st.session_state["nav"] = goto
 
        nav = st.radio(
            "Navigation",
            ["Upload", "Results", "About"],
            key="nav",
            label_visibility="collapsed",
        )
 
        st.markdown(
            '<div class="side-card">'
            '<div class="t">Quantitative Finance Tool</div>'
            'Markowitz Efficient Frontier<br>&amp; Michaud Resampling'
            '<br><br>© 2025 PortOpt<br>All rights reserved.'
            '<br><span class="vpill">v1.0.0</span>'
            '</div>'
            '<div class="side-row"><span>🌙&nbsp; Dark Mode</span><span class="tgl"></span></div>'
            f'<div class="side-row"><span>System Status</span>'
            f'<span class="ok-txt">Online</span></div>',
            unsafe_allow_html=True,
        )
    return nav
 
 
def _pagina_upload():
    _header_pagina(
        "Portfolio Optimizer",
        "Optimize your investment portfolio using Markowitz Efficient Frontier "
        "and Michaud Resampling.",
    )
 
    # ── DROP ZONE ───────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload Excel file",
        type=["xlsx", "xlsm"],
        key="uploader",
        label_visibility="collapsed",
    )
    st.markdown(
        '<div class="dz-hint">Supports .xlsx and .xlsm files &nbsp;•&nbsp; Max size 200MB</div>',
        unsafe_allow_html=True,
    )
 
    if uploaded is not None:
        if st.session_state["_last_file"] != uploaded.name:
            st.session_state.results = None
            st.session_state["_last_file"] = uploaded.name
 
        size_kb  = len(uploaded.getvalue()) / 1024
        size_str = f"{size_kb:.1f} KB" if size_kb < 1024 else f"{size_kb/1024:.2f} MB"
        st.markdown(
            f'<div class="file-card" style="margin-top:14px;">'
            f'<span style="font-size:1.5rem;">📄</span>'
            f'<div>'
            f'<div class="file-name">{uploaded.name}</div>'
            f'<div class="file-meta">{size_str} &nbsp;·&nbsp; Excel workbook</div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="info-banner">'
            '<div class="ib-ico"></div>'
            '<div>'
            '<div class="l1">Your file will be used to generate optimized portfolios.</div>'
            '<div class="l2">Make sure your file contains asset returns data in the correct format.</div>'
            '</div></div>',
            unsafe_allow_html=True,
        )
 
    # ── RUN ─────────────────────────────────────────────────────────────────
    st.markdown('<div style="height:14px;"></div>', unsafe_allow_html=True)
    btn_col, status_col = st.columns([1, 3])
    with btn_col:
        run_clicked = st.button("Run Optimization", disabled=(uploaded is None))
    with status_col:
        _status_slot = st.empty()
 
    if uploaded is None:
        _status_slot.markdown(_spill("Waiting for file…"), unsafe_allow_html=True)
    elif st.session_state.results is None and not run_clicked:
        _status_slot.markdown(_spill("Ready to optimize", "ok"), unsafe_allow_html=True)
    elif st.session_state.results is not None and not run_clicked:
        _status_slot.markdown(_spill("Optimization completed successfully", "ok"), unsafe_allow_html=True)
 
    # ── RUNNING ─────────────────────────────────────────────────────────────
    if run_clicked and uploaded is not None:
        _prog = st.progress(0, text="Starting…")
 
        def _set_status(msg):
            _status_slot.markdown(_spill(msg, "active"), unsafe_allow_html=True)
 
        def _set_progress(f):
            _prog.progress(min(f, 0.99), text=f"Michaud resampling: {int(f * 100)}%")
 
        suffix = ".xlsm" if uploaded.name.endswith(".xlsm") else ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name
 
        try:
            pf, pm, na = run_optimizer(
                tmp_path,
                progress_callback=_set_progress,
                status_callback=_set_status,
            )
            _prog.progress(1.0, text="Completed!")
            with open(tmp_path, "rb") as f:
                output_bytes = f.read()
            st.session_state.results = {
                "pf": pf, "pm": pm, "na": na,
                "bytes": output_bytes, "fname": uploaded.name,
            }
            st.session_state["_nav_goto"] = "Results"
            st.rerun()
        except Exception as e:
            _prog.empty()
            st.error(f"Error during optimization: {e}")
            _status_slot.markdown(_spill("Error occurred", "err"), unsafe_allow_html=True)
            st.session_state.results = None
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
 
    # ── FEATURE ROW + HELP ──────────────────────────────────────────────────
    _svg = lambda body: (
        "url(\"data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' "
        "viewBox='0 0 24 24' fill='none' stroke='%234C7DFF' stroke-width='2' "
        f"stroke-linecap='round' stroke-linejoin='round'>{body}</svg>\")"
    )
    ico_target  = _svg("<circle cx='12' cy='12' r='9'/><circle cx='12' cy='12' r='4'/><circle cx='12' cy='12' r='1'/>")
    ico_refresh = _svg("<path d='M21 12a9 9 0 1 1-2.6-6.4'/><path d='M21 3v6h-6'/>")
    ico_chart   = _svg("<path d='M21.2 15.9A10 10 0 1 1 8 2.8'/><path d='M22 12A10 10 0 0 0 12 2v10z'/>")
    ico_shield  = _svg("<path d='M12 22s8-3 8-10V5l-8-3-8 3v7c0 7 8 10 8 10z'/><path d='m9 12 2 2 4-4'/>")
 
    st.markdown(
        '<div class="feat-row">'
        f'<div class="feat"><div class="feat-ico" style="background-image:{ico_target};"></div>'
        '<div><div class="t">Mean-Variance Optimization</div><div class="s">Markowitz Efficient Frontier</div></div></div>'
        f'<div class="feat"><div class="feat-ico" style="background-image:{ico_refresh};"></div>'
        '<div><div class="t">Robust Resampling</div><div class="s">Michaud Block-Bootstrap</div></div></div>'
        f'<div class="feat"><div class="feat-ico" style="background-image:{ico_chart};"></div>'
        '<div><div class="t">Professional Results</div><div class="s">Charts, Tables &amp; Allocations</div></div></div>'
        f'<div class="feat"><div class="feat-ico" style="background-image:{ico_shield};"></div>'
        '<div><div class="t">Institutional Grade</div><div class="s">Reliable • Robust • Efficient</div></div></div>'
        '</div>'
        '<a class="help-pill" href="mailto:alessiobisceglia04@gmail.com">❔&nbsp; Need help?</a>',
        unsafe_allow_html=True,
    )
 
 
def _pagina_results():
    if not st.session_state.results:
        _header_pagina("Results", "Efficient frontiers, portfolio tables and allocations.")
        st.markdown(
            '<div class="empty-card">'
            '<div class="big">📊</div>'
            '<div class="title">No results yet</div>'
            'Upload an Excel file and run the optimization to see the efficient frontiers here.'
            '</div>',
            unsafe_allow_html=True,
        )
        return
 
    r  = st.session_state.results
    pf = r["pf"];  pm = r["pm"];  na = r["na"]
 
    _header_pagina("Results", r["fname"])
 
    # Metric cards
    st.markdown('<div class="sec-label">Results Overview</div>', unsafe_allow_html=True)
    all_p = pf + pm
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: st.markdown(_mcard("Best Expected Return", f"{max(p['rendimento'] for p in all_p):.2%}", accent=True), unsafe_allow_html=True)
    with c2: st.markdown(_mcard("Lowest Volatility",    f"{min(p['volatilita'] for p in all_p):.2%}"),              unsafe_allow_html=True)
    with c3: st.markdown(_mcard("Best Sharpe Ratio",    f"{max(p['sharpe'] for p in all_p):.4f}", accent=True),     unsafe_allow_html=True)
    with c4: st.markdown(_mcard("Number of Assets",     str(len(na))),                                              unsafe_allow_html=True)
    with c5: st.markdown(_mcard("Michaud Simulations",  "500"),                                                     unsafe_allow_html=True)
 
    st.divider()
 
    # Main comparison chart
    st.markdown('<div class="sec-label">Efficient Frontier Comparison</div>', unsafe_allow_html=True)
    st.plotly_chart(_build_frontier_chart(pf, pm), use_container_width=True)
 
    exp1, exp2 = st.columns(2)
    with exp1:
        with st.expander("View Markowitz Frontier"):
            st.plotly_chart(_build_single_chart(pf, C_MK, "Markowitz"), use_container_width=True)
    with exp2:
        with st.expander("View Michaud Frontier"):
            st.plotly_chart(_build_single_chart(pm, C_MI, "Michaud", line_shape="spline"), use_container_width=True)
 
    st.divider()
 
    # Comparison cards
    st.markdown('<div class="sec-label">Best Portfolio Comparison — Max Sharpe</div>', unsafe_allow_html=True)
    st.markdown(_compare_cards(pf, pm), unsafe_allow_html=True)
 
    st.divider()
 
    # Portfolio tables
    st.markdown('<div class="sec-label">Portfolio Tables</div>', unsafe_allow_html=True)
    tab_mk, tab_mi = st.tabs(["Markowitz Portfolios", "Michaud Portfolios"])
 
    idx_s_mk = max(range(len(pf)), key=lambda i: pf[i]["sharpe"])
    idx_v_mk = min(range(len(pf)), key=lambda i: pf[i]["volatilita"])
    idx_s_mi = max(range(len(pm)), key=lambda i: pm[i]["sharpe"])
    idx_v_mi = min(range(len(pm)), key=lambda i: pm[i]["volatilita"])
 
    with tab_mk:
        st.markdown(
            f"<small style='color:#9A9AA3;'>🟦 Max Sharpe — P{idx_s_mk+1} &nbsp;|&nbsp; ⬜ Min Volatility — P{idx_v_mk+1}</small>",
            unsafe_allow_html=True,
        )
        st.dataframe(_styled_table(pf, na, idx_s_mk, idx_v_mk), use_container_width=True, hide_index=True)
 
    with tab_mi:
        st.markdown(
            f"<small style='color:#9A9AA3;'>🟦 Max Sharpe — P{idx_s_mi+1} &nbsp;|&nbsp; ⬜ Min Volatility — P{idx_v_mi+1}</small>",
            unsafe_allow_html=True,
        )
        st.dataframe(_styled_table(pm, na, idx_s_mi, idx_v_mi), use_container_width=True, hide_index=True)
 
    st.divider()
 
    # Asset allocation comparison
    st.markdown('<div class="sec-label">Asset Allocation Comparison</div>', unsafe_allow_html=True)
 
    sel_col, mini_col = st.columns([1, 4])
    with sel_col:
        sel_p = st.selectbox("Portfolio", list(range(1, 11)), format_func=lambda x: f"Portfolio {x}")
    idx = sel_p - 1
    pesi_mk = pf[idx]["pesi"]
    pesi_mi = pm[idx]["pesi"]
 
    with mini_col:
        m1, m2, m3, m4 = st.columns(4)
        with m1: st.markdown(_mcard("MK Eff. Assets",   f"{_effective_n(pesi_mk):.1f}"),            unsafe_allow_html=True)
        with m2: st.markdown(_mcard("MI Eff. Assets",   f"{_effective_n(pesi_mi):.1f}"),            unsafe_allow_html=True)
        with m3: st.markdown(_mcard("MK Pos. > 5%",     str(sum(1 for p in pesi_mk if p > 0.05))), unsafe_allow_html=True)
        with m4: st.markdown(_mcard("MI Pos. > 5%",     str(sum(1 for p in pesi_mi if p > 0.05))), unsafe_allow_html=True)
 
    st.plotly_chart(_build_alloc_chart(pf, pm, na, idx), use_container_width=True)
 
    st.divider()
 
    # Download
    st.markdown('<div class="sec-label">Download</div>', unsafe_allow_html=True)
    _, dl_col, _ = st.columns([1, 2, 1])
    with dl_col:
        st.markdown(
            '<p style="text-align:center;color:#9A9AA3;font-size:0.85rem;margin-bottom:12px;">'
            'The Excel file has been updated with Markowitz and Michaud results.</p>',
            unsafe_allow_html=True,
        )
        st.download_button(
            label="Download Optimized Excel File",
            data=r["bytes"],
            file_name=r["fname"],
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            use_container_width=True,
        )
 
 
def _pagina_about():
    _header_pagina("About", "How the optimizer works and what the results mean.")
 
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
**Markowitz Mean-Variance Optimization**
Finds portfolio weights that minimize risk for each level of expected return,
tracing the classic efficient frontier via SLSQP optimization.
 
**Michaud Resampled Efficiency**
Runs 500 Monte Carlo simulations with block bootstrap resampling of historical
returns, then averages the resulting frontiers for more robust and
diversified allocations.
        """)
    with c2:
        st.markdown("""
**Technical details**
- 5 to 18 assets supported (dynamic)
- 500 Michaud simulations (block size adapted to data frequency)
- Data frequency auto-detected (daily / weekly / monthly) for annualization
- SLSQP optimizer via scipy
 
**Disclaimer**
Results are for educational and research purposes only.
This tool does not constitute financial advice.
        """)
 
 
# =============================================================================
# STREAMLIT — MAIN
# =============================================================================
 
def main():
    st.set_page_config(
        page_title="Portfolio Optimizer",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
 
    # Protezione globale: qualsiasi eccezione non catturata viene mostrata
    # nella pagina invece di lasciare lo schermo bianco.
    try:
        _main_body()
    except Exception as _fatal:
        st.error(f"Errore fatale durante il caricamento dell'app: {_fatal}")
        st.exception(_fatal)
 
 
def _main_body():
    st.markdown(CSS, unsafe_allow_html=True)
    st.markdown(CSS_PORTOPT, unsafe_allow_html=True)
 
    if "results" not in st.session_state:
        st.session_state.results = None
    if "_last_file" not in st.session_state:
        st.session_state["_last_file"] = None
 
    nav = _sidebar()
 
    if nav == "Upload":
        _pagina_upload()
    elif nav == "Results":
        _pagina_results()
    else:
        _pagina_about()
 
 
if __name__ == "__main__":
    main()
 
# Alessio Bisceglia
